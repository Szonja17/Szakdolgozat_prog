import requests
from bs4 import BeautifulSoup
import openpyxl

# Az oldal URL-je
url = 'https://potravinydomov.itesco.sk/groceries/en-GB/shop/bakery/all'

# HTTP GET kérés küldése az oldalhoz
response = requests.get(url)

# Ellenőrizzük a válasz státusz kódját
if response.status_code == 200:
    # Az oldal tartalmának betöltése a Beautiful Soup segítségével
    soup = BeautifulSoup(response.text, 'html.parser')

    # Az Excel fájl létrehozása és munkalap inicializálása
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Bakery Products"

    # A termékek nevének és árának keresése és Excel fájlba írása
    product_elements = soup.find_all('div', class_='product-details--content')
    price_elements = soup.find_all('span', class_='value')

    if len(product_elements) == len(price_elements):
        # Az adatok írása az Excel fájlba
        for i in range(len(product_elements)):
            product_name = product_elements[i].text.strip()
            product_price = price_elements[i].text.strip()
            worksheet.cell(row=i + 1, column=1, value=product_name)
            worksheet.cell(row=i + 1, column=2, value=product_price)

        # Az Excel fájl mentése
        workbook.save("bakery_products.xlsx")
        print("Az adatok sikeresen mentve az 'bakery_products.xlsx' fájlba.")
    else:
        print("Hiba: Nem sikerült megtalálni az összes termék nevét és árát.")
else:
    print("Hiba: Az oldal nem érhető el vagy hibát adott vissza.")